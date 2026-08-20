import hashlib
import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from app.services.catalog_import import CatalogRow


PLACEHOLDERS = {"", "-", "--", "n/a", "na", "none", "null", "not available", "tbd"}
PRODUCT_NAME_ALIASES = (
    "product_name",
    "name",
    "part_desc",
    "product_description",
    "description",
    "item_description",
)
PART_NUMBER_ALIASES = ("mfg_part_num", "manufacturer_part_number", "part_number", "mpn", "sku", "item_number")
BRAND_ALIASES = ("unilog_brand", "e1_brand", "dib_brand", "brand", "manufacturer", "part_manuf")


@dataclass(frozen=True)
class WorkbookRow:
    sheet_name: str
    row_number: int
    values: dict[str, Any]


@dataclass(frozen=True)
class WorkbookSheet:
    name: str
    header_row: int
    columns: list[str]
    original_columns: dict[str, str]
    rows: list[WorkbookRow]


@dataclass(frozen=True)
class ParsedWorkbook:
    filename: str
    content_sha256: str
    sheets: list[WorkbookSheet]

    @property
    def row_count(self) -> int:
        return sum(len(sheet.rows) for sheet in self.sheets)


class MultiSheetXLSXParser:
    """Parse every visible worksheet while preserving headers, worksheet names, and row lineage."""

    def __init__(self, max_rows: int = 500_000, header_scan_rows: int = 25):
        self.max_rows = max_rows
        self.header_scan_rows = header_scan_rows

    def parse(self, content: bytes, filename: str) -> ParsedWorkbook:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("The uploaded file is not a readable XLSX workbook") from exc
        sheets: list[WorkbookSheet] = []
        total_rows = 0
        try:
            for worksheet in workbook.worksheets:
                if worksheet.sheet_state != "visible":
                    continue
                parsed = self._parse_sheet(worksheet)
                if parsed is None:
                    continue
                total_rows += len(parsed.rows)
                if total_rows > self.max_rows:
                    raise ValueError(f"Workbook exceeds the {self.max_rows} data-row limit")
                sheets.append(parsed)
        finally:
            workbook.close()
        if not sheets:
            raise ValueError("Workbook contains no importable worksheets")
        return ParsedWorkbook(filename, hashlib.sha256(content).hexdigest(), sheets)

    def _parse_sheet(self, worksheet: Any) -> WorkbookSheet | None:
        buffered: list[tuple[int, tuple[Any, ...]]] = []
        iterator = worksheet.iter_rows(values_only=True)
        for row_number, values in enumerate(iterator, start=1):
            buffered.append((row_number, values))
            if row_number >= self.header_scan_rows:
                break
        if not buffered:
            return None
        header_row, raw_headers = max(buffered, key=lambda item: self._header_score(item[1]))
        if self._header_score(raw_headers) <= 0:
            return None
        columns = self._columns(raw_headers)
        original_columns = {
            column: str(raw_headers[index]).strip() if index < len(raw_headers) and raw_headers[index] not in (None, "") else column
            for index, column in enumerate(columns)
        }
        rows: list[WorkbookRow] = []
        all_rows = [*buffered, *[(number, values) for number, values in enumerate(iterator, start=len(buffered) + 1)]]
        for row_number, values in all_rows:
            if row_number <= header_row:
                continue
            payload = {
                column: self._cell(values[index] if index < len(values) else None)
                for index, column in enumerate(columns)
                if column and index < len(values) and self._cell(values[index]) is not None
            }
            if payload:
                rows.append(WorkbookRow(worksheet.title, row_number, payload))
        return WorkbookSheet(
            worksheet.title,
            header_row,
            [column for column in columns if column],
            original_columns,
            rows,
        )

    @staticmethod
    def _header_score(values: tuple[Any, ...]) -> int:
        strings = [value for value in values if isinstance(value, str) and value.strip()]
        if not strings:
            return 0
        identifiers = sum(bool(re.search(r"[A-Za-z]", value)) for value in strings)
        uniqueness = len({MultiSheetXLSXParser.normalize_header(value) for value in strings})
        return identifiers * 3 + uniqueness

    @classmethod
    def _columns(cls, values: tuple[Any, ...]) -> list[str]:
        columns: list[str] = []
        counts: dict[str, int] = {}
        for index, value in enumerate(values, start=1):
            base = cls.normalize_header(str(value)) if value not in (None, "") else f"column_{index}"
            counts[base] = counts.get(base, 0) + 1
            columns.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
        while columns and columns[-1].startswith("column_"):
            columns.pop()
        return columns

    @staticmethod
    def normalize_header(value: str) -> str:
        normalized = re.sub(r"[^a-zA-Z0-9]+", "_", value.strip()).strip("_").lower()
        return normalized or "column"

    @staticmethod
    def _cell(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if isinstance(value, str):
            stripped = value.strip()
            return None if stripped.lower() in PLACEHOLDERS else stripped
        return value


class XLSXCatalogImporter:
    def __init__(self, max_rows: int):
        self.parser = MultiSheetXLSXParser(max_rows=max_rows)

    def parse(self, content: bytes, filename: str, sheet_name: str | None = None) -> list[CatalogRow]:
        workbook = self.parser.parse(content, filename)
        selected = [sheet for sheet in workbook.sheets if sheet_name is None or sheet.name == sheet_name]
        if not selected:
            raise ValueError(f"Worksheet '{sheet_name}' was not found")
        rows: list[CatalogRow] = []
        for sheet in selected:
            if sheet.name.lower().replace("_", " ").strip() == "delivery format" and sheet_name is None:
                continue
            for item in sheet.rows:
                product_name = self._product_name(item.values)
                if not product_name:
                    continue
                raw_content = "\n".join(
                    f"{key.replace('_', ' ').title()}: {self._display(value)}"
                    for key, value in item.values.items()
                    if value not in (None, "", [])
                )
                rows.append(
                    CatalogRow(
                        product_name=product_name,
                        source_identifier=f"{filename}:{sheet.name}:row-{item.row_number}",
                        raw_content=raw_content,
                        row_number=item.row_number,
                    )
                )
        if not rows:
            raise ValueError("Workbook contains no importable catalog rows")
        return rows

    @staticmethod
    def _product_name(values: dict[str, Any]) -> str:
        for alias in PRODUCT_NAME_ALIASES:
            if values.get(alias):
                description = str(values[alias]).strip()
                part = next((str(values[key]).strip() for key in PART_NUMBER_ALIASES if values.get(key)), "")
                brand = next((str(values[key]).strip() for key in BRAND_ALIASES if values.get(key)), "")
                return " ".join(dict.fromkeys(piece for piece in (brand, part, description) if piece))[:255]
        part = next((str(values[key]).strip() for key in PART_NUMBER_ALIASES if values.get(key)), "")
        brand = next((str(values[key]).strip() for key in BRAND_ALIASES if values.get(key)), "")
        return " ".join(piece for piece in (brand, part) if piece)[:255]

    @staticmethod
    def _display(value: Any) -> str:
        if isinstance(value, (list, dict)):
            return str(value)
        return str(value)
