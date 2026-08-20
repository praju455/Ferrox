from io import BytesIO

from openpyxl import Workbook

from app.models import ReferenceDataset
from app.services.reference_data import ReferenceDataService


def make_workbook(headers, rows, title="Sheet1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_reference_loader_versions_manufacturers_and_matches_names(db_session):
    service = ReferenceDataService(db_session, max_rows=100)
    first = service.load_xlsx(
        "manufacturer",
        make_workbook(
            ["Unilog Brand", "E1 Brand", "Manufacturer Website"],
            [["Bombas Boyser", "Boyser Pumps", "https://www.boyser.com/products"]],
        ),
        "manufacturers.xlsx",
    )
    match = service.canonical_manufacturer("Bombas Boyser")

    assert first.status == "ready"
    assert match and match["value"] == "Bombas Boyser"
    assert service.manufacturer_domains("Bombas Boyser") == {"boyser.com"}

    second = service.load_xlsx(
        "manufacturer",
        make_workbook(["Unilog Brand"], [["Acme Industrial"]]),
        "manufacturers-v2.xlsx",
    )
    db_session.refresh(first)
    assert not first.is_active
    assert second.is_active
    assert db_session.query(ReferenceDataset).count() == 2


def test_reference_loader_reads_lov_uom_and_side_by_side_fractions(db_session):
    service = ReferenceDataService(db_session, max_rows=100)
    service.load_xlsx(
        "lov",
        make_workbook(
            ["Attribute Name", "LOV Value"],
            [["Material", "Brass"], ["Material", "Stainless Steel"], ["Finish", "Chrome"]],
        ),
        "lov.xlsx",
    )
    service.load_xlsx("uom", make_workbook(["UOM", "Description"], [["EA", "Each"]]), "uom.xlsx")
    service.load_xlsx(
        "fraction",
        make_workbook(
            ["Fraction", "Decimal", "Fraction", "Decimal"],
            [["1/8", 0.125, "1/2", 0.5], ["1/4", 0.25, "3/4", 0.75]],
        ),
        "fractions.xlsx",
    )

    assert service.allowed_values(attribute="Material") == {"Brass", "Stainless Steel"}
    assert service.canonical_uom("EA") == "EA"
    assert service.fraction_for_decimal(0.5) == "1/2"


def test_ground_truth_metadata_preserves_all_252_delivery_headers(db_session):
    workbook = Workbook()
    workbook.active.title = "Input"
    workbook.active.append(["Mfg Part Num", "Part Desc"])
    workbook.active.append(["P-1", "Pump"])
    delivery = workbook.create_sheet("Delivery Format")
    headers = ["Manufacturer Part Number", *[f"Delivery Field {index}" for index in range(2, 253)]]
    delivery.append(headers)
    delivery.append(["P-1", *[f"value-{index}" for index in range(2, 253)]])
    buffer = BytesIO()
    workbook.save(buffer)

    dataset = ReferenceDataService(db_session).load_xlsx("ground_truth", buffer.getvalue(), "truth.xlsx")

    assert len(dataset.columns["Delivery Format"]) == 252
    assert dataset.dataset_metadata["delivery_columns"] == headers
