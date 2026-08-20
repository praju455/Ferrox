from io import BytesIO

from openpyxl import Workbook

from app.models import Product, ProductDeliveryRecord
from app.services.evaluation import GroundTruthEvaluationService
from app.services.reference_data import ReferenceDataService


def ground_truth() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Input"
    workbook.active.append(["SKU", "Part Desc"])
    workbook.active.append(["SKU-1", "BRASS FITTING"])
    workbook.active.append(["SKU-2", "STEEL FITTING"])
    output = workbook.create_sheet("Delivery Format")
    output.append(
        [
            "SKU",
            "Manufacturer Name",
            "Invoice Desc",
            "Mobile Desc",
            "Classpath",
            "Material",
        ]
    )
    mobile = "Acme Manufacturing ACME Fitting Series SKU-1 Brass Chrome Finish"
    output.append(["SKU-1", "Acme Manufacturing", "BRASS FITTING SKU-1", mobile, "Plumbing > Fittings", "Brass"])
    output.append(["SKU-2", "Acme Manufacturing", "STEEL FITTING SKU-2", mobile, "Plumbing > Fittings", "Steel"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_evaluation_scores_fields_limits_manufacturer_taxonomy_and_unmatched_rows(db_session):
    references = ReferenceDataService(db_session)
    dataset = references.load_xlsx("ground_truth", ground_truth(), "ground-truth.xlsx")
    lov = Workbook()
    lov.active.append(["Attribute Name", "LOV Value"])
    lov.active.append(["Material", "Brass"])
    lov.active.append(["Material", "Steel"])
    stream = BytesIO()
    lov.save(stream)
    references.load_xlsx("lov", stream.getvalue(), "lov.xlsx")
    product = Product(name="SKU-1 Fitting", category="Plumbing > Fittings")
    db_session.add(product)
    db_session.flush()
    mobile = "Acme Manufacturing ACME Fitting Series SKU-1 Brass Chrome Finish"
    db_session.add(
        ProductDeliveryRecord(
            product_id=product.id,
            fields={
                "SKU": "SKU-1",
                "Manufacturer Name": "Acme Manufacturing",
                "Invoice Desc": "BRASS FITTING SKU-1",
                "Mobile Desc": mobile,
                "Classpath": "Plumbing > Fittings",
                "Material": "Brass",
            },
            descriptions={},
            quality={},
        )
    )
    db_session.commit()

    run = GroundTruthEvaluationService(db_session).run(dataset.id, generate_missing=False)

    assert run.total_items == 2
    assert run.matched_items == 1
    assert run.field_accuracy == 0.5
    assert run.character_limit_compliance == 0.5
    assert run.lov_compliance == 0.5
    assert run.manufacturer_accuracy == 0.5
    assert run.taxonomy_accuracy == 0.5
    assert run.row_results[1]["matched"] is False
    assert "SKU-1" in GroundTruthEvaluationService(db_session).export_csv(run)
