from io import BytesIO

from openpyxl import Workbook


def workbook_file(headers, rows, sheet_name="Sheet1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def ground_truth_file() -> bytes:
    workbook = Workbook()
    input_sheet = workbook.active
    input_sheet.title = "Input"
    input_sheet.append(["Mfg Part Num", "Part Desc"])
    input_sheet.append(["F-100", "BRASS ELBOW"])
    delivery = workbook.create_sheet("Delivery Format")
    headers = [
        "Manufacturer Part Number",
        "Manufacturer Name",
        "Product Title",
        "Invoice Desc",
        "Mobile Desc",
        *[f"Delivery Field {index}" for index in range(6, 253)],
    ]
    delivery.append(headers)
    delivery.append(["F-100", "Acme Industrial", "ACME F-100 ELBOW", "ELBOW F-100", "Acme Industrial F-100 Elbow Brass Plumbing Fitting Product", *([""] * 247)])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_xlsx_catalog_import_queues_rows_from_multiple_input_sheets(client):
    workbook = Workbook()
    first = workbook.active
    first.title = "Pumps"
    first.append(["Mfg Part Num", "Part Desc", "Brand"])
    first.append(["P-1", "Pump", "Acme"])
    second = workbook.create_sheet("Fittings")
    second.append(["Mfg Part Num", "Part Desc", "Brand"])
    second.append(["F-1", "Fitting", "Acme"])
    stream = BytesIO()
    workbook.save(stream)

    response = client.post(
        "/api/v1/imports/catalog",
        files={"file": ("catalog.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 202
    assert response.json()["imported_rows"] == 2
    assert {item["payload"]["catalog_row_number"] for item in response.json()["items"]} == {2}


def test_reference_delivery_and_evaluation_api_contract(client):
    manufacturer = workbook_file(
        ["MANUFACTURER_NAME", "BRAND_NAME", "Manufacturer Website"],
        [["Acme Industrial", "ACME®", "https://acme.example/products"]],
    )
    upload = client.post(
        "/api/v1/reference-data/manufacturer",
        files={"file": ("manufacturer.xlsx", manufacturer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 201
    assert upload.json()["row_count"] == 1

    truth = client.post(
        "/api/v1/reference-data/ground_truth",
        files={"file": ("ground-truth.xlsx", ground_truth_file(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert truth.status_code == 201
    assert len(truth.json()["dataset_metadata"]["delivery_columns"]) == 252

    product = client.post("/api/v1/products", json={"name": "Acme F-100 Elbow"}).json()
    client.patch(f"/api/v1/products/{product['id']}/fields/manufacturer", json={"value": "Acme Industrial"})
    client.patch(f"/api/v1/products/{product['id']}/fields/manufacturer_part_number", json={"value": "F-100"})
    delivery = client.post(f"/api/v1/products/{product['id']}/delivery")

    assert delivery.status_code == 200
    assert len(delivery.json()["fields"]) == 252
    assert delivery.json()["fields"]["Manufacturer Name"] == "Acme Industrial"
    assert delivery.json()["quality"]["schema_ready"] is True

    evaluation = client.post(
        "/api/v1/evaluations",
        json={"ground_truth_dataset_id": truth.json()["id"], "generate_missing_deliveries": False},
    )
    assert evaluation.status_code == 201
    assert evaluation.json()["total_items"] == 1
    assert evaluation.json()["matched_items"] == 1
    report = client.get(f"/api/v1/evaluations/{evaluation.json()['id']}/report.csv")
    assert report.status_code == 200
    assert "F-100" in report.text
