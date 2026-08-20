from io import BytesIO

from openpyxl import Workbook

from app.models import ExtractedField, Product
from app.services.delivery import UnilogDeliveryService
from app.services.description_builder import DeterministicDescriptionBuilder
from app.services.reference_data import ReferenceDataService


def xlsx(headers, rows, title="Sheet1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_deterministic_descriptions_obey_order_casing_and_limits():
    builder = DeterministicDescriptionBuilder(fraction_lookup=lambda value: "1/4" if value == 0.25 else None)
    result = builder.build(
        "Fitting",
        {
            "manufacturer": "Acme Manufacturing",
            "brand": "ACME®",
            "series": "Pro Series",
            "manufacturer_part_number": "F-100",
            "product_type": "Elbow Fitting",
            "fitting_type": "90 Degree Elbow",
            "material": "Stainless Steel",
        },
    )

    assert result.product_title.startswith("ACME® Pro Series F-100 Elbow Fitting")
    assert result.invoice_description == result.invoice_description.upper()
    assert len(result.invoice_description) <= 40
    assert len(result.mobile_description) <= 80
    assert result.long_description == (
        "ACME®, Elbow Fitting, Pro Series, F-100, 90 Degree Elbow, Stainless Steel"
    )
    assert builder.format_measurement(50.25, "in") == "50-1/4 in"


def test_delivery_generation_uses_exact_252_headers_and_master_manufacturer(db_session):
    references = ReferenceDataService(db_session)
    references.load_xlsx(
        "manufacturer",
        xlsx(
            ["MANUFACTURER_NAME", "BRAND_NAME", "E1 Brand"],
            [["Rheem Manufacturing", "FRIGIDAIRE®", "Frigidaire"]],
        ),
        "manufacturer.xlsx",
    )
    workbook = Workbook()
    workbook.active.title = "Input"
    workbook.active.append(["Mfg Part Num", "Part Desc"])
    workbook.active.append(["PDSH4816AF", "Dishwasher"])
    delivery_sheet = workbook.create_sheet("Delivery Format")
    headers = [
        "Manufacturer Name",
        "Brand Name",
        "Manufacturer Part Number",
        "Product Title",
        "Invoice Desc",
        "Mobile Desc",
        *[f"Customer Field {index}" for index in range(7, 253)],
    ]
    delivery_sheet.append(headers)
    delivery_sheet.append(["Rheem Manufacturing", "FRIGIDAIRE®", "PDSH4816AF", *([""] * 249)])
    stream = BytesIO()
    workbook.save(stream)
    references.load_xlsx("ground_truth", stream.getvalue(), "ground-truth.xlsx")

    product = Product(name="PDSH4816AF Dishwasher", category="Dishwasher")
    db_session.add(product)
    db_session.flush()
    db_session.add_all(
        [
            ExtractedField(
                product_id=product.id,
                field_name="manufacturer",
                value="Frigidaire",
                unit=None,
                confidence=0.9,
                alternatives=[],
            ),
            ExtractedField(
                product_id=product.id,
                field_name="manufacturer_part_number",
                value="PDSH4816AF",
                unit=None,
                confidence=0.95,
                alternatives=[],
            ),
        ]
    )
    db_session.commit()

    record = UnilogDeliveryService(db_session).generate(product)

    assert len(record.fields) == 252
    assert list(record.fields) == headers
    assert record.fields["Manufacturer Name"] == "Rheem Manufacturing"
    assert record.fields["Brand Name"] == "FRIGIDAIRE®"
    assert record.fields["Manufacturer Part Number"] == "PDSH4816AF"
    assert record.fields["Customer Field 252"] == ""
    assert record.quality["schema_ready"] is True
    assert record.quality["unsupported_fields_left_blank"] is True
